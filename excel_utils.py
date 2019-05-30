import openpyxl
from datetime import datetime

class Excel_worker():

    def __init__(self, excel_path, excel_sheet, excel_headers_row,  num_object_excel_col_name, update_object_excel_col_name, parsing_date_excel_col_name, html_excel_col_name):
        self.excel_path = excel_path  # Путь к файлу ексель для парсинга
        self.wb = openpyxl.load_workbook(filename=excel_path)  # открываем существующий ексель
        # DEPRECATED self.ws = self.wb.get_sheet_by_name(excel_sheet)  # Выбираем лист
        self.ws = self.wb[excel_sheet]  # Выбираем лист
        self.excel_headers_row = excel_headers_row  # Номер строки с заголовками столбцов
        self.num_object_excel_col_name = num_object_excel_col_name  # Название столбца с номером объявления
        self.update_object_excel_col_name = update_object_excel_col_name  # Название столбца с датой обновления
        self.parsing_date_excel_col_name = parsing_date_excel_col_name  # Название столбца с датой парсинга
        self.html_excel_col_name = html_excel_col_name  # Название столбца, в который записывается путь к файлу ХТМЛ (гиперссылка на файл)


    def get_column_by_name(self, column_name):  # получить столбец ("А") по названию поля, Название поля = "Номер объявления"
        for cell in self.ws[self.excel_headers_row]:
            if cell.value == column_name:
                return cell.column

    def get_present_object_list(self):  # Получаем словарь с записями в ексель ПЕРЕД началом парсинга {(1206333, '15.12.2017'): 4, (1035344, '15.12.2017'): 5, (1174760, '15.12.2017'): 6, ...} Возвращает словарь, где ключ - кортеж из номера объявления и строки с датой обновления, а значение - номер строки данного объявления
        num_object_column = self.get_column_by_name(self.num_object_excel_col_name)  # Столбец с номерами объявлений
        update_object_column = self.get_column_by_name(self.update_object_excel_col_name)  # Столбец с датами обновлений
        return {(self.ws.cell(row=row, column=num_object_column).value, self.ws.cell(row=row, column=update_object_column).value): row for row in range(self.excel_headers_row + 1, self.ws.max_row + 1)}  # {(1206333, '15.12.2017') : 1, ...} проходимся по каждой строке и извлекаем номер объявления, дату обновления и номер строки

    def write_into_cell(self, project, row_num):  # Записать в ячейку Екселя, где project = {название поля: значение, ...} - сформированный с реалта
        for key in project:  # проходимся по каждому полю в проекте
            column_number = self.get_column_by_name(key)   # получаем столбец по названию поля
            if key == self.html_excel_col_name:  # если название поля=="Ссылка на HTML", то нужно записать гиперссылку на файл
                cell=self.ws.cell(row=row_num, column=column_number)
                cell.hyperlink = project[key]
            else:  # для остальных полей просто записываем значение
                cell = self.ws.cell(row=row_num, column=column_number)
                cell.value = project[key]

    def add_projects_into_existing_excel(self, projects): # записывает все projects (лист со словарями) в ексель, т.е. все объявления со страницы
        row_num = self.ws.max_row # находим последнюю строку в ексель(чтобы записывать новые данные в следующую)
        for project in projects:   # проходимся по кадому проекту
            row_num +=1
            self.write_into_cell(project, row_num)
        self.wb.save(self.excel_path)

    def calculate_expozition(self, num_object):  # рассчитываем экспозицию для объекта(завершенного), где num_object - номер объявления
        update_list = []  # для обэявления нужно найти минимульную дату обновления - она будет начальной датой
        for cell in self.ws[self.get_column_by_name(self.num_object_excel_col_name)]:  # проходимся по всем ячейкам столбца с номерами объявлений
            if num_object == cell.value:  # если номер объявления искомый равен ячейке
                cell_name = self.ws.cell(row=cell.row, column=self.get_column_by_name(self.update_object_excel_col_name))
                str_update = cell_name.value  # то извлекаем дату обновления
                object_update = datetime.strptime(str_update, "%d.%m.%Y")  # переводим дату обновления из тип "строки" в тип "даты"
                update_list.append(object_update)  # и добавляем в список дат
        update = min(update_list)  # находим минималную дату обновления
        expozition = datetime.now() - update  # рассчитываем экспозицию как сегодняшняя дата минус минимальная дата обновления
        return expozition.days  # возвращаем экспозицию в днях

    def rewrite_parsing_date(self, all_projects_in_excel, todays_date):  # Заменяем дату парсинга для объявлений с реалта, которые существуют в екселе (с той же датой обновления)
        for project in all_projects_in_excel:  # all_projects_in_excel = [(3, 123, '01.01.2018')...], где 3 - номер строки
            print("rewrite parsing date FOR: ", project)
            cell = self.ws.cell(row=project[0], column=self.get_column_by_name(self.parsing_date_excel_col_name))
            cell.value = todays_date
            self.wb.save(self.excel_path)

    def write_expozition(self, excel_objects, id_object_name_at_pages, expozition_col_name, todays_date):
        id_excel_objects = [key[0] for key in excel_objects]  # Получаем список номеров объявлений, которые есть в Ексель
        difference = set(id_excel_objects) - set(id_object_name_at_pages)  # разница, для которой нужно записать расчетную экспозицию
        id_row_excel_objects = {key[0]: excel_objects[key] for key in excel_objects}
        new_difference = [object for object in difference if self.ws["%s%d" % (self.get_column_by_name(expozition_col_name),id_row_excel_objects[object])].value is None]
        for object in new_difference:
            print("write compozition for closed object and change their parsing date FOR: ", object)
            expozition = self.calculate_expozition(object)
            cell_expozition = self.ws.cell(row=id_row_excel_objects[object], column=self.get_column_by_name(expozition_col_name))
            cell_expozition.value = expozition
            cell_todays_day = self.ws.cell(row=id_row_excel_objects[object], column=self.get_column_by_name(self.parsing_date_excel_col_name))
            cell_todays_day.value = todays_date
            self.wb.save(self.excel_path)








