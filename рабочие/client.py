import openpyxl
from datetime import datetime

num_object_name = "№ Объявления"
update_object_name = "Дата обновления"
parsing_date = "Дата парсинга"
headers_row = 3
headers_dict = {}

colnames = ['Дата актуальности предложения', '№ Объявления']

excel_path = "Форма_04_Предложения_продажи_и_аренды_NEW.xlsx"
excel_sheet = "Предложения"
column_num = 70 # Количество столбов в екселе


wb = openpyxl.load_workbook(filename=excel_path)# открываем существующий ексель
ws = wb.get_sheet_by_name(excel_sheet) # Выбираем лист


def get_present_object_list(ws, num_object_name, update_object_name, headers_row):
    return [(ws["%s%d"%(get_column_by_name(ws, num_object_name, headers_row), row)].value, ws["%s%d"%(get_column_by_name(ws, update_object_name, headers_row), row)].value) for row in range(headers_row+1, ws.max_row)]

def get_present_object_list1(ws, num_object_name, update_object_name, headers_row):  # return [(1206333, '15.12.2017'), (1035344, '15.12.2017'), (1174760, '15.12.2017'), ...]
    num_object_column = get_column_by_name(ws, num_object_name)
    print(num_object_column)
    update_object_column = get_column_by_name(ws, update_object_name)
    print(update_object_column)
    # for row in range(headers_row+1, ws.max_row):
    return {(ws["%s%d"%(num_object_column, row)].value, ws["%s%d"%(update_object_column, row)].value): row for row in range(headers_row+1, ws.max_row+1)}
    # return {(ws["%s%d"%(num_object_column, row)].value, ws["%s%d"%(update_object_column, row)].value): row for row in range(headers_row+1, ws.max_row)}


def get_column_by_name(ws, column_name, headers_row=3):
    for cell in ws[headers_row]:
        if cell.value == column_name:
            return cell.column


# def get_present_object_list_1(ws, num_object_column, update_object_column, headers_row):num_object_column = get_column_by_name(ws, num_object_name, headers_row)
# num_object_column = get_column_by_name(ws, num_object_name, headers_row)
# update_object_column = get_column_by_name(ws, update_object_name, headers_row)
# print([(ws["%s%d"%(num_object_column, row)].value, ws["%s%d"%(update_object_column, row)].value) for row in range(headers_row+1, ws.max_row)])
    # return [(ws[])]


# a = get_column_by_name(ws, "Дата обновления")
a = get_present_object_list1(ws, num_object_name, update_object_name, headers_row)
print(a)
# for column in ws["C"]:
#     print(column.value)
# for row in ws.rows[4]:
#     print(row.value)
#
# for column in ws[3]:
#     print(column.value)


# def iter_rows(ws):
# for row in ws.iter_rows():
#     print(row)
#     for i in row:
#         print(i.value)
    # yield [cell.value for cell in row]

# rows = iter_rows(ws)
# print(rows)

row_num = ws.max_row # находим последнюю строку (чтобы записывать новые данные в следующую)
# for row in ws['C1''F100']:
#     for cell in row:
#         print(cell.value)
#     print(i.value)
# for i in ws.columns:
#     print(i)