import openpyxl
from bs4 import BeautifulSoup
import requests
from datetime import datetime, date
my_set = {1, 2}
excel_path = "Форма_04_Предложения_продажи_и_аренды_NEW.xlsx"
excel_sheet = "Предложения"

wb = openpyxl.load_workbook(filename=excel_path)  # открываем существующий ексель
ws = wb.get_sheet_by_name(excel_sheet)


def get_column_by_name(column_name):
    for cell in ws[3]:
        if cell.value == column_name:
            return cell.column

def get_html(url):
    try:
        res = requests.get(url, {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/59.0.3071.115 Safari/537.36'})
    except requests.ConnectionError:
        return
    if res.status_code < 400:
        return res.content

# for num_object in list(my_set):
#     print('№ объявления на поиск: ', num_object)
#     update_list = []
#     for cell in ws['C']:
#         print(cell.value)
#         if num_object == cell.value:
#             print('Yes')
#             str_update = ws["%s%d" % (get_column_by_name("Дата обновления"), cell.row)].value
#             print(str_update)
#             object_update = datetime.strptime(str_update, "%d.%m.%Y")
#             print(object_update)
#             update_list.append(object_update)
#
#     update = min(update_list)
#     print(update)
#     expozition = datetime.now() - update
#     print(expozition.days)

baseurl = 'https://realt.by/sale/shops/object/1159139/'

html_obj = get_html(baseurl)
soup = BeautifulSoup(html_obj, "html.parser")
table = soup.find('div', {'class': 'text-12 mb20 fl wp100'})
print(table.text)
# for i in table:
#     print(i)

